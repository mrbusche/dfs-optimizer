# https://www.youtube.com/watch?v=zibV6xaOGEA

import os

import openpyxl
import pandas as pd
from pulp import LpMaximize, LpProblem, LpVariable, lpSum

POSITION = "DK Position"
PROJECTION = "DK Projection"
SALARY = "DK Salary"
PLAYER = "Player"

salaries = {}
points = {}

salary_cap = 50000
max_rows = 15
four_wr = {"QB": 1, "RB": 2, "WR": 4, "TE": 1, "DST": 1}
three_rb = {"QB": 1, "RB": 3, "WR": 3, "TE": 1, "DST": 1}
two_te = {"QB": 1, "RB": 2, "WR": 3, "TE": 2, "DST": 1}

players = pd.read_csv(
    r"draftkings.csv",
    usecols=[PLAYER, POSITION, PROJECTION, SALARY],
)

wb = openpyxl.Workbook()
ws = wb.active

available_players = players.groupby([POSITION, PLAYER, PROJECTION, SALARY]).agg("count")
available_players = available_players.reset_index()

for pos in available_players[POSITION].unique():
    available_pos = available_players[available_players[POSITION] == pos]
    salaries[pos] = list(
        available_pos[[PLAYER, SALARY]].set_index(PLAYER).to_dict().values()
    )[0]
    points[pos] = list(
        available_pos[[PLAYER, PROJECTION]].set_index(PLAYER).to_dict().values()
    )[0]


def calculate(lineup_type, file_name):
    total_score = 0
    for lineup in range(1, max_rows + 1):
        _vars = {k: LpVariable.dict(k, v, cat="Binary") for k, v in points.items()}

        prob = LpProblem("Fantasy", LpMaximize)
        rewards = []
        costs = []

        for k, v in _vars.items():
            costs += lpSum([salaries[k][i] * _vars[k][i] for i in v])
            rewards += lpSum([points[k][i] * _vars[k][i] for i in v])
            prob += lpSum([_vars[k][i] for i in v]) == lineup_type[k]

        prob += lpSum(rewards)
        prob += lpSum(costs) <= salary_cap
        if lineup != 1:
            prob += lpSum(rewards) <= total_score - 0.001
        prob.solve()

        score = str(prob.objective)
        column_number = 1

        for v in prob.variables():
            score = score.replace(v.name, str(v.varValue))
            if v.varValue != 0:
                ws.cell(row=lineup, column=column_number).value = v.name
                column_number += 1

        total_score = eval(score)
        ws.cell(row=lineup, column=column_number).value = total_score

    wb.save(file_name + ".xlsx")
    pd.read_excel(file_name + ".xlsx").to_csv(file_name + ".csv", index=False)


calculate(three_rb, "three_rb")
calculate(four_wr, "four_wr")
calculate(two_te, "two_te")

path = os.getcwd()
file_list = [path + "\\four_wr.csv", path + "\\three_rb.csv", path + "\\two_te.csv"]
print("file_list", file_list)
excl_list = []

for idx, file in enumerate(file_list):
    excl_list.append(pd.read_csv(file))  # , header=0 if idx == 0 else None
